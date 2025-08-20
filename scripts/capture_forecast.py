# scripts/capture_forecast.py
import os, requests, pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook

API_KEY = os.environ["WEATHER_API_KEY"]  # <- set in repo secrets
airports = ["BOM", "DEL", "CCU", "MAA", "GOI", "COK", "AMD", "LHR", "LGW",
            "SYD", "MEL", "HKG", "HND", "ICN", "FRA", "PNQ", "HKT", "BKK"]

dfs = []
for code in airports:
    url = (
        "https://api.weather.com/v3/wx/forecast/hourly/2day"
        f"?iataCode={code}&units=e&language=en-US&format=json&apiKey={API_KEY}"
    )
    resp = requests.get(url, timeout=60)
    resp.raise_for_status()
    data = resp.json()
    df = pd.DataFrame(data)

    # Parse timestamps when present
    if 'validTimeUtc' in df.columns:
        df['validTimeUtc'] = pd.to_datetime(df['validTimeUtc'], unit='s', utc=True)
    if 'validTimeLocal' in df.columns:
        df['validTimeLocal'] = pd.to_datetime(df['validTimeLocal'])

    df['airport'] = code
    dfs.append(df)

df_all = pd.concat(dfs, ignore_index=True)

desired_cols = [
    'airport',
    'validTimeLocal', 'validTimeUtc',
    'pressureMeanSeaLevel', 'relativeHumidity', 'qpf',
    'temperature', 'temperatureDewPoint', 'temperatureFeelsLike',
    'visibility', 'windDirection', 'windGust', 'windSpeed',
    'uvIndex', 'iconCode', 'iconCodeExtended'
]
available_cols = [c for c in desired_cols if c in df_all.columns]
final_df = df_all[available_cols].copy()

# Pretty datetime strings for Excel
for col in ("validTimeLocal", "validTimeUtc"):
    if col in final_df.columns:
        final_df[col] = pd.to_datetime(final_df[col], errors='coerce').dt.strftime("%Y-%m-%d %H:%M")

# Sheet named like "18-8-25"
today = datetime.now()
sheet_name = f"{today.day}-{today.month}-{today.strftime('%y')}"

XLSX_PATH = "weather_forecast_capture.xlsx"

# Create (or open) workbook and replace today's sheet
if os.path.exists(XLSX_PATH):
    wb = load_workbook(XLSX_PATH)
else:
    wb = Workbook()
    wb.remove(wb.active)  # drop default empty sheet

if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    wb.remove(ws)

ws = wb.create_sheet(title=sheet_name)
ws.append(list(final_df.columns))
for row in final_df.itertuples(index=False):
    ws.append(list(row))

wb.save(XLSX_PATH)
print(f"Saved {len(final_df)} rows to {XLSX_PATH} [{sheet_name}]")
